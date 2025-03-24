#!/usr/bin/env python
# coding: utf-8

# # Industrial ecology open online course (IEooc)
# ## IEooc_Methods3_Software1: Dynamic stock modelling
# 
# Overview of how basic dynamic stock modelling with ODYM, the Python framework for dynamic MFA, works. First, the installation of the dynamic stock model functionw within ODYM is explained, and then a combination of inflow-driven and stock-driven modelling is demonstraded. For this tutorial the steel stock of China is estimated from historic data on final steel consumption (1900-2008) and then extrapolated with an exogenous saturation scenario (2009-2100)
# 
# For more info:
# https://github.com/IndEcol/ODYM 

# ### 1) Dynamic Stock Modelling in Python - Documentation and Tutorial
# 

# Dynamic models of material and product stocks require standard routines for computation. The dynamic_stock_model library offers a toolbox for computation and modification of dynamic stock models, including inflow-driven (van der Voet 2002) and stock-driven (Müller 2006) modelling.
# Below you find a quick tutorial and demonstration of the central feastures of the dynamic stock model functions in ODYM.

# ### Dependencies

# Python 3.0 or later <br>
#     numpy >= 1.9 <br>
#     scipy >= 0.14 <br>

# ### Installation and import

# #### Manual installation, by modifying the Python path

# The ODYM MFA software, of which the current dynamic stock model class is part of, is not available as a package yet. Instead, you can directly download/save the following GitHub file: <br>
# https://github.com/IndEcol/ODYM/blob/master/odym/modules/dynamic_stock_model.py  <br>
# on your machine, and add the location of that file to your system path:

# In[1]:


import sys
sys.path.append('C:\\Users\\[YOUR PATH]\\ODYM-RECC\\ODYM_Model\\odym\\modules') 


# Just put your own path in the command above, and use \\ for subfolders. Again, the paths needs to point to the folder where the \_\_init\_\_.py is located.
# Now, the class can be imported:

# In[2]:


import dynamic_stock_model as dsm # import the dynamic stock model library


# And other libraries later needed can be imported:

# In[3]:


import openpyxl
import numpy as np
import matplotlib.pyplot as plt

# For Ipython Notebook only
get_ipython().run_line_magic('matplotlib', 'inline')


# Now, the class is ready to use. 
# 
# ### Tutorial for simple inflow-driven model and stock-driven model
# 
# Below a combination of inflow-driven and stock-driven modelling is demonstraded. For this tutorial the steel stock of China is estimated from historic data on final steel consumption (1900-2008) and then extrapolated with an exogenous saturation scenario (2009-2100)
# 
# ### 1. Load and check data

# In[4]:


DSM_Datafile  = openpyxl.load_workbook(filename='IEooc_Methods3_Software1_Data.xlsx')
DSM_Datasheet = DSM_Datafile['Data_Steel_China']

Historic_Years = [] #years 1900 to 2008
Historic_Inflow = [] # apparent final steel consumption, China, kt/yr
Future_Years = [] # years 2009-2100
Future_Stock = [] # future scenario for the steel stock, China, kt

for m in range(3,112): # Read historic data
    Historic_Years.append(int(DSM_Datasheet.cell(m,3).value))
    Historic_Inflow.append(DSM_Datasheet.cell(m,4).value)

for m in range(3,94): # Read future scenario
    Future_Years.append(int(DSM_Datasheet.cell(m,9).value))
    Future_Stock.append(DSM_Datasheet.cell(m,10).value)

print(Historic_Years)

print(Future_Stock)


# ### 2. Inflow-driven model 1900-2008
# With these data we can now define a dynamic stock model by creating an instance of the DynamicStockModel class and assigning time, inflow, and lifetime. For the average product lifetime we chose a normally distributed lifetime with the mean of 30 years and a standard deviation of 30% of the mean.

# In[5]:


AvgLifetime = 30
InputStdDev = 0.3*AvgLifetime

China_DSM = dsm.DynamicStockModel(t = np.array(Historic_Years), i = np.array(Historic_Inflow), lt = {'Type': 'Normal', 'Mean': np.array([AvgLifetime]), 'StdDev': np.array([InputStdDev]) })


# Let's check the attributes of the object we just defined:<br>
# t indicates the time in years <br>

# In[6]:


China_DSM.t


# i indicates the Inputs <br>

# In[7]:


China_DSM.i


# lt indicates the lifetime <br>

# In[8]:


China_DSM.lt


# Regarding the attribute lt, you can see that the class automatically expands the lifetime distribution parameters from scalar to vector. <br>
# At any time, we can determine a checkstring that describes the content of the TestDSM object:

# In[9]:


CheckStr = China_DSM.dimension_check()
print(CheckStr)


# The check string comes in simple html format to be included in html logfiles. <br>
# Now, we compute the stock s(t,t') of age-cohort t' in year t as
# $$s(t,t') = i(t')\cdot(1-\sum_{t''=0}^{t-t'}\lambda(t'')) $$
# Here, $\lambda(t'')$ is the probability function of discard at age $t''$. It is determined using the lifetime distribution type specified above. The full model is docomented here: http://pubs.acs.org/doi/suppl/10.1021/es201904c/suppl_file/es201904c_si_001.pdf <br>
# The result is the stock, broken down by age-cohort:

# In[10]:


Stock_by_cohort = China_DSM.compute_s_c_inflow_driven()
print(Stock_by_cohort)
print(Stock_by_cohort.shape)


# The dimension of the stock is 109x109. The row index is the model year, the column index is the age-cohort. <br>
# We continue by computing the total stock, the outflow by age-cohort, the total outflow, and the stock change: <br>
# 
# S indicates the total stock

# In[11]:


S   = China_DSM.compute_stock_total()
print(S)


# O_C indicates Outflow by Cohort <br>
# S_C indicates Stock by Cohort

# In[12]:


O_C = China_DSM.compute_o_c_from_s_c()
print(O_C)


# O indicates Outflow total

# In[13]:


O   = China_DSM.compute_outflow_total()
print(O)


# DS indicates Stock Changes (Delta Stock) 

# In[14]:


DS  = China_DSM.compute_stock_change()
print(DS)


# Now the dynamic stock is fully determined: Both stock and outflow are broken down by age-cohort and as total. We can make a final check by printing the dimension check and by computing the stock balance:

# In[15]:


print(China_DSM.dimension_check()) # dimension_check returns two variables, but we only print the first one, which has index 0.


# Bal indicates the Stock Balance

# In[16]:


Bal = China_DSM.check_stock_balance()
print(Bal)


# Let's make a few plots to see what happened:

# In[17]:


plt.imshow(China_DSM.s_c, interpolation='nearest')
plt.xlabel('age-cohort')
plt.ylabel('year')
plt.title('Stock by age-cohort')
plt.show();


# In the upper plot the upper triangle is always empty, as there are no age-cohorts from the future present in the stock. Since the stock is growing rapidly, the youngest age-cohorts dominate, and these are shown in the lower right corner of the plot.

# In[18]:


plt.imshow(China_DSM.o_c,interpolation='nearest')
plt.xlabel('age-cohort')
plt.ylabel('year')
plt.title('Outflow by age-cohort')
plt.show();


# Above the age-cohorts that contribute most to the outflows around 2008 are those between age-cohorts 80 and 90 (1980-1990)
# 
# ### 3. stock-driven model (2009-2100)

# In stock-driven modelling, only the stock is known at first, and we want to know the inflow, the outflow, and the cohort structure of both stock and outflow. The DynamicStockModel class can do that. 
# 
# Here we not only have to solve the stock-driven model for future years but also take into account the stock as a result of historic consumption 1900-2008. There are several ways to do that (https://github.com/stefanpauliuk/dynamic_stock_model), and here we take a more elegant approach: We apply stock-driven modelling to the entire time series, starting from 1900, and concatenating the historic result and the future scenario. As the stock-driven model is the inverse operation of the flow-driven model the resulting inflow for the time 1900-2008 will be identical to the supplied input data provided that the exact same lifetime distribution is used.

# In[19]:


China_Stock_Full = np.concatenate((S,Future_Stock), axis =0)
Time_Full        = np.concatenate((np.array(Historic_Years),np.array(Future_Years)), axis =0)

China_DSM_Full = dsm.DynamicStockModel(t = Time_Full, s = China_Stock_Full, lt = {'Type': 'Normal', 'Mean': np.array([AvgLifetime]), 'StdDev': np.array([0.3*AvgLifetime]) })
CheckStr = China_DSM_Full.dimension_check()
print(CheckStr)


# As above, we now call a sequence of methods to find all the missing information in the stock model. The central method for stock-driven modelling is __compute_stock_driven_model()__, which contains a recursive calculation starting from the fist year to determine outflow, inflow, and stock by cohorts using the lifetime model and the mass balance. The model is documented here:
# http://pubs.acs.org/doi/suppl/10.1021/es201904c/suppl_file/es201904c_si_001.pdf

# S_C indicates Stock by Cohort <br>
# O_C indicates Outflow by Cohort <br>
# I indicates Inflows <br>
# O indicates Total Outflow <br>
# DS indicates Stock Changes (Delta Stock) <br>
# Bal indicates the Stock Balance <br>

# In[20]:


S_C, O_C, I = China_DSM_Full.compute_stock_driven_model()
O   = China_DSM_Full.compute_outflow_total()
DS  = China_DSM_Full.compute_stock_change()
Bal = China_DSM_Full.check_stock_balance()
print(Bal)


# Well, let's have a look at the result!
# 
# Below we plot stocks and flows on one axis by looking at the annual quantity represented by the flows (kt/yr --> kt for one year).

# In[21]:


plt1, = plt.plot(China_DSM_Full.t, China_DSM_Full.i)
plt2, = plt.plot(China_DSM_Full.t, China_DSM_Full.s)
plt3, = plt.plot(China_DSM_Full.t, China_DSM_Full.o)
plt4, = plt.plot([2008,2008],[0,2e7], color = 'k', linestyle = '--')
plt.xlabel('Year')
plt.ylabel('tons')
plt.title('Stock parameters')
plt.legend([plt1,plt2,plt3], ['Inflow','Stock','Outflow'], loc = 2)
plt.show();


# We can see that the inflow of steel peaks around 2020, which is around the inflection point of the exogenous stock curve. The ditch in the inflow in 2009 is not a consequence of the economic crisis at that time but a result of the slight mismatch between the historic time series and the future scenario, as both were taken from different versions of the database when building this exercise. The outflow of steel peaks later, shortly after the steek stock itself will have peaked around 2040.

# In[22]:


plt.imshow(China_DSM_Full.s_c,interpolation='nearest')
plt.xlabel('age-cohort')
plt.ylabel('year')
plt.title('Stock by age-cohort')
plt.show();


# In[23]:


plt.imshow(China_DSM_Full.o_c,interpolation='nearest')
plt.xlabel('age-cohort')
plt.ylabel('year')
plt.title('Outflow by age-cohort')
plt.show();


# We also need to check whether the model was able to reprodue the historic inflow:

# In[24]:


FlowBal = China_DSM_Full.i[0:109] - Historic_Inflow
print(FlowBal)
print(np.abs(FlowBal).sum())


# This calculation confirms that the model works correctly. With the same lifetime distribution the historic apparent steel consumption is reproduced, and the difference between the two inflow time series (the original historic data and the re-calculated one from the stock-driven model) is practically zero.